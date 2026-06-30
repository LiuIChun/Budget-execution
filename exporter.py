import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

def export_execution_report(summary_df, expense_df, output_name="budget_execution_report.xlsx"):
    """
    Export budget execution report to Excel with multiple sheets and formatting.
    
    Args:
        summary_df (DataFrame): Summary data with department execution info
        expense_df (DataFrame): Detailed expense data
        output_name (str): Output filename
    
    Returns:
        Path: Path to the exported Excel file
    """
    output_path = Path("output") / output_name
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Prepare data for each sheet
    # Sheet1: 各系所 (Department details)
    sheet1_df = summary_df.copy()
    # Ensure we have required columns
    if 'department_name' not in sheet1_df.columns:
        sheet1_df['department_name'] = sheet1_df.get('department', 'Unknown')
    if 'budget_amount' not in sheet1_df.columns:
        sheet1_df['budget_amount'] = 0
    if 'actual_amount' not in sheet1_df.columns:
        sheet1_df['actual_amount'] = 0
    if 'execution_rate' not in sheet1_df.columns:
        sheet1_df['execution_rate'] = 0
    
    # Sheet2: 學院統計 (College statistics)
    if 'college' in summary_df.columns:
        college_stats = summary_df.groupby('college').agg({
            'budget_amount': 'sum',
            'actual_amount': 'sum'
        }).reset_index()
        college_stats['execution_rate'] = (
            college_stats['actual_amount'] / college_stats['budget_amount'] * 100
        ).fillna(0)
        sheet2_df = college_stats
    else:
        sheet2_df = pd.DataFrame({
            'college': ['無學院分類'],
            'budget_amount': [summary_df['budget_amount'].sum() if 'budget_amount' in summary_df.columns else 0],
            'actual_amount': [summary_df['actual_amount'].sum() if 'actual_amount' in summary_df.columns else 0],
            'execution_rate': [0]
        })
    
    # Sheet3: 排行 (Ranking)
    sheet3_df = summary_df.copy()
    if 'execution_rate' in sheet3_df.columns:
        sheet3_df = sheet3_df.sort_values('execution_rate', ascending=False)
    else:
        sheet3_df['execution_rate'] = 0
        sheet3_df = sheet3_df.sort_values('execution_rate', ascending=False)
    
    # Sheet4: 購案明細 (Purchase details)
    sheet4_df = expense_df.copy()
    # Ensure expected columns exist
    expected_cols = ['item_name', 'amount', 'date', 'category', 'department_name']
    for col in expected_cols:
        if col not in sheet4_df.columns:
            sheet4_df[col] = ''
    
    # Write to Excel with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write each sheet
        sheet1_df.to_excel(writer, sheet_name='各系所', index=False)
        sheet2_df.to_excel(writer, sheet_name='學院統計', index=False)
        sheet3_df.to_excel(writer, sheet_name='排行', index=False)
        sheet4_df.to_excel(writer, sheet_name='購案明細', index=False)
        
        # Get the workbook and apply formatting
        workbook = writer.book
        
        # Define styles
        title_font = Font(size=14, bold=True, color="00008B")  # Dark blue
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue
        percent_format = '0.00%'  # Percentage with 2 decimal places
        thousand_format = '#,##0'  # Thousands separator
        
        # Apply formatting to each sheet
        for sheet_name in ['各系所', '學院統計', '排行', '購案明細']:
            worksheet = writer.sheets[sheet_name]
            
            # Freeze first row
            worksheet.freeze_panes = 'A2'
            
            # Format header row
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply number formatting based on column names
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    col_name = worksheet.cell(row=1, column=cell.column).value
                    if col_name in ['execution_rate', '執行率(%)', '執行率']:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = percent_format
                    elif col_name in ['budget_amount', 'actual_amount', '金額', 'amount', '預算金額', '執行金額']:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = thousand_format
                    elif col_name in ['日期', 'date']:
                        if isinstance(cell.value, str):
                            try:
                                # Try to convert to date format
                                cell.value = pd.to_datetime(cell.value)
                                cell.number_format = 'YYYY-MM-DD'
                            except:
                                pass
        
        # Add conditional formatting to execution rate columns (if exists)
        for sheet_name in ['各系所', '學院統計', '排行']:
            worksheet = writer.sheets[sheet_name]
            # Find execution rate column
            exec_rate_col = None
            for col in range(1, worksheet.max_column + 1):
                header_val = worksheet.cell(row=1, column=col).value
                if header_val in ['execution_rate', '執行率(%)', '執行率']:
                    exec_rate_col = get_column_letter(col)
                    break
            
            if exec_rate_col:
                # Apply color scale: red (low) to yellow to green (high)
                color_scale = ColorScaleRule(
                    start_type='percentile', start_value=0, start_color='F8696B',  # Red
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',     # Yellow
                    end_type='percentile', end_value=100, end_color='63BE7B'     # Green
                )
                worksheet.conditional_formatting.add(
                    f'{exec_rate_col}2:{exec_rate_col}{worksheet.max_row}',
                    color_scale
                )
    
    return output_path